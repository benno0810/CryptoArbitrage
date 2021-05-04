import time
import json

import openpyxl
from openpyxl import Workbook

try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

import traceback


def get_hb():

    with open('huobi.json') as f:
        txt = f.read()
        hb = json.loads(txt)
    return sorted(hb, key=lambda x: x['created-at'], reverse=False)


def get_binance():
    # excel中数据已经按datetime排序了
    wb = openpyxl.load_workbook('binance.xlsx')  # 打开excel文件
    sheet = wb['Sheet1']  # 获取工作表

    rows, cols = sheet.max_row, sheet.max_column

    alls = []
    for r in range(2, rows + 1):

        row = []
        for c in range(1, cols+1):
            row.append(sheet.cell(row=r, column=c).value)

        tmp = {
            'created-at': int((time.mktime(row[0].timetuple())) * 1000),
            'market': row[1],
            'type': row[2],
            'price': row[3],
            'filled-amount': row[4],
            'done_money': row[5],
            'fee': row[6],
            'coin': row[7]
        }
        alls.append(tmp)
    return alls


def get_only_binance():
    '''
    10198; 58500
    diff: 48302
    '''
    hb = get_hb()
    bi = get_binance()

    num_hb = len(hb)
    num_bi = len(bi)
    duichong = []
    next_begin = 0
    bi_begin = 0

    for j in range(bi_begin, num_bi):
        #print('j = ', j,  next_begin, num_hb)
        for i in range(next_begin, num_hb):
            #print(i, bi[j]['created-at'], hb[i]['created-at'])
            if bi[j]['created-at'] > hb[i]['created-at']:
                #print(bi[j]['filled-amount'], hb[i-1]['filled-amount'], hb[i]['filled-amount'])
                #if abs(bi[j]['filled-amount'] - float(hb[i]['filled-amount'])) < 0.00001:
                if abs(bi[j]['filled-amount'] - float(hb[i]['filled-amount'])) == 0:
                    duichong.append(bi[j])
                    next_begin = i + 1
                #elif abs(bi[j]['filled-amount'] - float(hb[i-1]['filled-amount'])) < 0.00001:
                elif abs(bi[j]['filled-amount'] - float(hb[i-1]['filled-amount'])) == 0:
                    duichong.append(bi[j])
                    next_begin = i
                else:
                    next_begin = i + 1
                break

    print(len(duichong))
    for one in duichong:
        bi.remove(one)

    print(len(bi))

    with open('only.json', 'w') as f:
        f.write(json.dumps(bi))

if __name__ == '__main__':
    get_only_binance()

